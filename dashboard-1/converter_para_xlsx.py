import sqlite3
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

def formatar_planilha(ws, df):
    """Formata a planilha com cabeçalho em negrito e ajusta largura das colunas"""
    # Formata o cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Ajusta a largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 40)

def converter_para_xlsx(banco_dados, arquivo_saida='dados_exportados.xlsx'):
    """
    Converte todas as tabelas de um banco de dados SQLite para um único arquivo XLSX,
    com cada tabela em uma aba diferente.
    
    Args:
        banco_dados (str): Caminho para o arquivo do banco de dados SQLite
        arquivo_saida (str): Nome do arquivo XLSX de saída
    """
    # Conecta ao banco de dados
    conn = sqlite3.connect(banco_dados)
    cursor = conn.cursor()
    
    # Cria um novo arquivo Excel
    wb = Workbook()
    
    # Remove a planilha padrão
    wb.remove(wb.active)
    
    # Obtém a lista de tabelas
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tabelas = [t[0] for t in cursor.fetchall()]
    
    if not tabelas:
        print("Nenhuma tabela encontrada no banco de dados.")
        return
    
    print(f"\nIniciando conversão de {len(tabelas)} tabelas para XLSX...\n")
    
    # Para cada tabela, exporta para uma aba do Excel
    for tabela_nome in tabelas:
        try:
            # Lê a tabela para um DataFrame do pandas
            df = pd.read_sql_query(f"SELECT * FROM \"{tabela_nome}\"", conn)
            
            # Cria uma nova aba com o nome da tabela (limitado a 31 caracteres, que é o limite do Excel)
            nome_aba = tabela_nome[:31]
            ws = wb.create_sheet(title=nome_aba)
            
            # Adiciona os dados à planilha
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Formata a planilha
            formatar_planilha(ws, df)
            
            print(f"✓ Tabela '{tabela_nome}' exportada para a aba '{nome_aba}'")
            
        except Exception as e:
            print(f"❌ Erro ao exportar tabela '{tabela_nome}': {str(e)}")
    
    # Salva o arquivo Excel
    wb.save(arquivo_saida)
    conn.close()
    
    print(f"\nConversão concluída! Arquivo salvo como '{arquivo_saida}'")

if __name__ == "__main__":
    banco_dados = "DBV Capital_Objetivos.db"
    arquivo_saida = "DBV_Capital_Objetivos.xlsx"
    converter_para_xlsx(banco_dados, arquivo_saida)
